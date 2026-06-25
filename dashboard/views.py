"""
Dashboard Views
Admin dashboard with live stats, Chart.js data, candidate management,
election configuration, results, export, and email notifications.
Includes auto-close fix for expired elections.
"""

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views import View
from django.http import HttpResponse
from django.contrib import messages
from django.utils import timezone
from django.conf import settings
from django.core.mail import send_mass_mail
from rest_framework.views import APIView
from rest_framework.response import Response

from elections.models import Election
from candidates.models import Candidate
from votes.models import Vote
from accounts.models import User, VoterProfile, AuditLog


class DashboardHomeView(LoginRequiredMixin, View):
    def get(self, request):
        if request.user.is_admin:
            return redirect('dashboard:admin_home')
        return redirect('dashboard:voter_home')


class AdminDashboardView(LoginRequiredMixin, View):
    template_name = 'dashboard/admin.html'

    def get(self, request):
        if not request.user.is_admin:
            return redirect('dashboard:voter_home')

        # ── Auto-close expired elections (FIX) ──────────────────
        Election.objects.filter(
            status='active',
            end_datetime__lt=timezone.now()
        ).update(status='closed')
        # ──────────────────────────────────────────────────────────

        elections = Election.objects.all().order_by('-created_at')
        active_election = elections.filter(status='active').first()
        total_voters = VoterProfile.objects.filter(is_eligible=True).count()
        total_voted = VoterProfile.objects.filter(has_voted=True).count()
        turnout_pct = round((total_voted / total_voters * 100), 1) if total_voters else 0
        recent_logs = AuditLog.objects.select_related('user').order_by('-timestamp')[:20]

        context = {
            'elections': elections,
            'active_election': active_election,
            'total_voters': total_voters,
            'total_voted': total_voted,
            'turnout_pct': turnout_pct,
            'recent_logs': recent_logs,
            'candidates': active_election.candidates.filter(is_active=True) if active_election else [],
        }
        return render(request, self.template_name, context)


class VoterDashboardView(LoginRequiredMixin, View):
    template_name = 'dashboard/voter.html'

    def get(self, request):
        # Auto-close expired elections here too
        Election.objects.filter(
            status='active', end_datetime__lt=timezone.now()
        ).update(status='closed')

        active_elections = Election.objects.filter(status='active')
        voter_votes = Vote.objects.filter(voter=request.user).values_list('election_id', flat=True)
        elections_data = []
        for election in active_elections:
            has_voted = election.id in voter_votes
            elections_data.append({
                'election': election,
                'has_voted': has_voted,
                'can_vote': election.is_active and not has_voted,
            })
        return render(request, self.template_name, {
            'elections_data': elections_data, 'voter_votes': voter_votes,
        })


class ElectionResultsView(LoginRequiredMixin, View):
    """Animated results page with winner announcement and live prediction."""
    template_name = 'elections/results.html'

    def get(self, request, election_id):
        election = get_object_or_404(Election, id=election_id)
        candidates = list(election.candidates.filter(is_active=True))
        total_votes = election.total_votes

        for c in candidates:
            c.vote_pct = round((c.vote_count / total_votes * 100), 1) if total_votes else 0

        candidates.sort(key=lambda x: x.vote_count, reverse=True)

        winner = candidates[0] if candidates and election.status in ['closed', 'results'] else None
        winner_pct = winner.vote_pct if winner else 0

        prediction = candidates[0] if election.is_active and candidates else None

        total_voters = VoterProfile.objects.filter(is_eligible=True).count()
        turnout_pct = round((total_votes / total_voters * 100), 1) if total_voters else 0

        return render(request, self.template_name, {
            'election': election, 'candidates': candidates, 'winner': winner,
            'winner_pct': winner_pct, 'prediction': prediction,
            'total_votes': total_votes, 'turnout_pct': turnout_pct,
        })


class ExportResultsView(LoginRequiredMixin, View):
    """Export election results to Excel or PDF."""

    def get(self, request, election_id):
        if not request.user.is_admin:
            return redirect('dashboard:home')

        election = get_object_or_404(Election, id=election_id)
        export_format = request.GET.get('format', 'excel')
        candidates = election.candidates.filter(is_active=True)
        total_votes = election.total_votes

        if export_format == 'excel':
            return self._export_excel(election, candidates, total_votes)
        return self._export_pdf(election, candidates, total_votes)

    def _export_excel(self, election, candidates, total_votes):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Election Results"

            header_fill = PatternFill("solid", fgColor="0D47A1")
            header_font = Font(bold=True, color="FFFFFF", size=12)

            ws.merge_cells('A1:E1')
            ws['A1'] = f"Election Results: {election.title}"
            ws['A1'].font = Font(bold=True, size=14, color="0D47A1")
            ws['A1'].alignment = Alignment(horizontal='center')
            ws['A2'] = f"Date: {election.end_datetime.strftime('%d %b %Y')}"
            ws['A3'] = f"Total Votes: {total_votes}"

            headers = ['Rank', 'Candidate Name', 'Party', 'Votes', 'Percentage']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=5, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')

            sorted_candidates = sorted(candidates, key=lambda c: c.vote_count, reverse=True)
            for rank, candidate in enumerate(sorted_candidates, 1):
                pct = round((candidate.vote_count / total_votes * 100), 1) if total_votes else 0
                ws.append([rank, candidate.name, candidate.party_name, candidate.vote_count, f"{pct}%"])

            ws.column_dimensions['A'].width = 8
            ws.column_dimensions['B'].width = 25
            ws.column_dimensions['C'].width = 25
            ws.column_dimensions['D'].width = 12
            ws.column_dimensions['E'].width = 12

            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="results_{election.id}.xlsx"'
            wb.save(response)
            return response
        except ImportError:
            messages.error(self.request if hasattr(self, 'request') else None, 'openpyxl not installed. Run: pip install openpyxl')
            return redirect('dashboard:admin_home')

    def _export_pdf(self, election, candidates, total_votes):
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.lib import colors
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            import io

            buffer = io.BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=A4)
            styles = getSampleStyleSheet()
            elements = []

            elements.append(Paragraph(f"<b>Election Results: {election.title}</b>", styles['Title']))
            elements.append(Spacer(1, 12))
            elements.append(Paragraph(f"Total Votes: {total_votes} | Date: {election.end_datetime.strftime('%d %b %Y')}", styles['Normal']))
            elements.append(Spacer(1, 20))

            data = [['Rank', 'Candidate', 'Party', 'Votes', '%']]
            sorted_candidates = sorted(candidates, key=lambda c: c.vote_count, reverse=True)
            for rank, c in enumerate(sorted_candidates, 1):
                pct = round((c.vote_count / total_votes * 100), 1) if total_votes else 0
                data.append([rank, c.name, c.party_name, c.vote_count, f"{pct}%"])

            table = Table(data, colWidths=[50, 150, 150, 80, 80])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0D47A1')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 10),
            ]))
            elements.append(table)
            doc.build(elements)
            buffer.seek(0)

            response = HttpResponse(buffer, content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="results_{election.id}.pdf"'
            return response
        except ImportError:
            return redirect('dashboard:admin_home')


class SendEmailNotificationView(LoginRequiredMixin, View):
    template_name = 'dashboard/send_email.html'

    def get(self, request):
        if not request.user.is_admin:
            return redirect('dashboard:home')
        elections = Election.objects.filter(status='active')
        return render(request, self.template_name, {'elections': elections})

    def post(self, request):
        if not request.user.is_admin:
            return redirect('dashboard:home')

        subject = request.POST.get('subject', '').strip()
        message_body = request.POST.get('message', '').strip()
        notify_type = request.POST.get('notify_type', 'all')

        if not subject or not message_body:
            messages.error(request, 'Subject and message are required.')
            return redirect('dashboard:send_email')

        if notify_type == 'not_voted':
            voters = User.objects.filter(
                role='voter', is_active=True,
                voter_profile__has_voted=False, voter_profile__is_eligible=True
            )
        else:
            voters = User.objects.filter(role='voter', is_active=True)

        emails = tuple(
            (subject, message_body, settings.DEFAULT_FROM_EMAIL, [v.email]) for v in voters
        )

        try:
            send_mass_mail(emails, fail_silently=False)
            AuditLog.log('admin_action', request, request.user, f'Email sent to {voters.count()} voters')
            messages.success(request, f'Email sent successfully to {voters.count()} voters!')
        except Exception as e:
            messages.error(request, f'Failed to send emails: {str(e)}')

        return redirect('dashboard:admin_home')


# ── REST API ─────────────────────────────────────────────────────

class TurnoutAPIView(APIView):
    def get(self, request, election_id=None):
        if not request.user.is_authenticated:
            return Response({'error': 'Unauthorized'}, status=401)

        if election_id:
            election = get_object_or_404(Election, id=election_id)
            total = VoterProfile.objects.filter(is_eligible=True).count()
            voted = Vote.objects.filter(election=election).count()
            candidates = election.candidates.filter(is_active=True)

            top = max(candidates, key=lambda c: c.vote_count) if candidates else None
            candidate_data = sorted([
                {
                    'name': c.name, 'party': c.party_name, 'votes': c.vote_count,
                    'pct': round(c.vote_count / voted * 100, 1) if voted else 0,
                    'is_leading': c == top,
                } for c in candidates
            ], key=lambda x: x['votes'], reverse=True)

            return Response({
                'election': election.title,
                'total_voters': total,
                'votes_cast': voted,
                'not_voted': total - voted,
                'turnout_percent': round(voted / total * 100, 1) if total else 0,
                'candidates': candidate_data,
                'leader': top.name if top and top.vote_count > 0 else None,
                'last_updated': timezone.now().isoformat(),
            })

        elections = Election.objects.all()
        return Response({'elections': [
            {'id': e.id, 'title': e.title, 'status': e.status, 'votes': e.total_votes} for e in elections
        ]})


class ElectionConfigView(LoginRequiredMixin, View):
    template_name = 'elections/configure.html'

    def get(self, request, election_id=None):
        if not request.user.is_admin:
            return redirect('dashboard:home')
        election = get_object_or_404(Election, id=election_id) if election_id else None
        return render(request, self.template_name, {'election': election})

    def post(self, request, election_id=None):
        if not request.user.is_admin:
            return redirect('dashboard:home')

        data = request.POST
        if election_id:
            election = get_object_or_404(Election, id=election_id)
            action = 'election_modified'
        else:
            election = Election()
            election.created_by = request.user
            action = 'election_created'

        title = data.get('title', '').strip()
        start_raw = data.get('start_datetime', '').strip()
        end_raw = data.get('end_datetime', '').strip()

        if not title:
            messages.error(request, 'Election title is required.')
            return render(request, self.template_name, {'election': election})

        if not start_raw or not end_raw:
            messages.error(request, 'Start and End date/time are required.')
            return render(request, self.template_name, {'election': election})

        try:
            from datetime import datetime
            import pytz

            start_dt = datetime.strptime(start_raw, '%Y-%m-%dT%H:%M')
            end_dt = datetime.strptime(end_raw, '%Y-%m-%dT%H:%M')

            if end_dt <= start_dt:
                messages.error(request, 'End date must be after Start date.')
                return render(request, self.template_name, {'election': election})

            tz = pytz.timezone('Asia/Kolkata')
            election.start_datetime = tz.localize(start_dt)
            election.end_datetime = tz.localize(end_dt)
            election.title = title
            election.description = data.get('description', '')
            election.instructions = data.get('instructions', '')
            election.status = data.get('status', 'draft')
            election.save()

            AuditLog.log(action, request, request.user, f"Election: {election.title}")
            messages.success(request, f'Election "{election.title}" saved successfully!')
            return redirect('dashboard:admin_home')

        except Exception as e:
            messages.error(request, f'Error saving election: {str(e)}')
            return render(request, self.template_name, {'election': election})
